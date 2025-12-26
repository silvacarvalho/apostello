"""
Serviço de Relatórios PDF/Excel
Exportação de escalas, participações e avaliações
"""
from typing import List, Optional
from datetime import date, datetime
from io import BytesIO
from sqlalchemy.orm import Session
import logging

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.escala import Escala, StatusEscala
from app.models.item_escala import ItemEscala
from app.models.usuario import Usuario
from app.models.igreja import Igreja
from app.models.avaliacao import Avaliacao
from app.models.distrito import Distrito

logger = logging.getLogger(__name__)


class RelatorioService:
    """Serviço para geração de relatórios"""

    def __init__(self, db: Session):
        self.db = db

    # ==================== ESCALA PDF ====================
    def gerar_escala_pdf(self, escala_id: int, igreja_id: Optional[int] = None) -> BytesIO:
        """
        Gera PDF da escala mensal.
        Se igreja_id for informado, gera apenas os cultos dessa igreja (formato compacto, uma página).
        """
        escala = self.db.query(Escala).filter(Escala.id == escala_id).first()
        if not escala:
            raise ValueError("Escala não encontrada")
        
        distrito = self.db.query(Distrito).filter(Distrito.id == escala.distrito_id).first()
        
        # Filtrar itens
        query = self.db.query(ItemEscala).filter(ItemEscala.escala_id == escala_id)
        if igreja_id:
            query = query.filter(ItemEscala.igreja_id == igreja_id)
            # Se filtrado por igreja, ordenar por data/horário
            itens = query.order_by(ItemEscala.data_culto, ItemEscala.horario).all()
        else:
            # Se geral, ordenar por igreja primeiro, depois data/horário
            itens = query.order_by(ItemEscala.igreja_id, ItemEscala.data_culto, ItemEscala.horario).all()
        
        # Buscar igreja se filtrado
        igreja_filtro = None
        if igreja_id:
            igreja_filtro = self.db.query(Igreja).filter(Igreja.id == igreja_id).first()

        buffer = BytesIO()
        
        # Se for por igreja, usar formato retrato (cabe em uma página)
        if igreja_id:
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=A4,
                leftMargin=1.5*cm,
                rightMargin=1.5*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
        else:
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=landscape(A4),
                leftMargin=1*cm,
                rightMargin=1*cm,
                topMargin=1*cm,
                bottomMargin=1*cm
            )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16 if igreja_id else 18,
            alignment=TA_CENTER,
            spaceAfter=8 if igreja_id else 12
        )
        
        if igreja_id and igreja_filtro:
            elements.append(Paragraph(
                f"Escala de Pregação e Louvor",
                title_style
            ))
            # Nome da igreja em destaque
            igreja_style = ParagraphStyle(
                'IgrejaTitulo',
                parent=styles['Heading2'],
                fontSize=14,
                alignment=TA_CENTER,
                spaceAfter=6
            )
            elements.append(Paragraph(
                f"{igreja_filtro.nome}",
                igreja_style
            ))
        else:
            elements.append(Paragraph(
                f"Escala de Pregação e Louvor - {escala.mes:02d}/{escala.ano}",
                title_style
            ))
        
        # Subtítulo com distrito e mês/ano
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11 if igreja_id else 12,
            alignment=TA_CENTER,
            spaceAfter=15 if igreja_id else 20
        )
        
        # Traduzir status para português
        status_map = {
            'RASCUNHO': 'Rascunho',
            'PUBLICADA': 'Publicada',
            'ARQUIVADA': 'Arquivada'
        }
        status_texto = status_map.get(escala.status.value if hasattr(escala.status, 'value') else str(escala.status), str(escala.status))
        
        if igreja_id:
            elements.append(Paragraph(
                f"{escala.mes:02d}/{escala.ano} - Distrito: {distrito.nome if distrito else 'N/A'} - Status: {status_texto}",
                subtitle_style
            ))
        else:
            elements.append(Paragraph(
                f"Distrito: {distrito.nome if distrito else 'N/A'} - Status: {status_texto}",
                subtitle_style
            ))
        
        # Tabela de dados - colunas diferentes se filtrado por igreja
        if igreja_id:
            # Sem coluna de igreja (já está no título)
            data = [["Data", "Dia", "Horário", "Pregador", "Cantor", "Tema"]]
        else:
            data = [["Data", "Dia", "Horário", "Igreja", "Pregador", "Cantor", "Tema"]]
        
        # Mapeamento de dias da semana em português
        dias_semana_pt = {
            0: "Seg",
            1: "Ter",
            2: "Qua",
            3: "Qui",
            4: "Sex",
            5: "Sáb",
            6: "Dom"
        }
        
        for item in itens:
            pregador = self.db.query(Usuario).filter(Usuario.id == item.pregador_id).first() if item.pregador_id else None
            cantor = self.db.query(Usuario).filter(Usuario.id == item.cantor_id).first() if item.cantor_id else None
            
            dia_semana = dias_semana_pt.get(item.data_culto.weekday(), "") if item.data_culto else ""
            
            if igreja_id:
                data.append([
                    item.data_culto.strftime("%d/%m") if item.data_culto else "",
                    dia_semana,
                    item.horario.strftime("%H:%M") if item.horario else "",
                    pregador.nome_completo if pregador else "A definir",
                    cantor.nome_completo if cantor else "-",
                    item.tema_customizado or (item.tema.titulo if item.tema else "-")
                ])
            else:
                igreja = self.db.query(Igreja).filter(Igreja.id == item.igreja_id).first()
                data.append([
                    item.data_culto.strftime("%d/%m") if item.data_culto else "",
                    dia_semana,
                    item.horario.strftime("%H:%M") if item.horario else "",
                    igreja.nome if igreja else "",
                    pregador.nome_completo if pregador else "A definir",
                    cantor.nome_completo if cantor else "-",
                    item.tema_customizado or (item.tema.titulo if item.tema else "-")
                ])
        
        # Criar tabela
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Corpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (2, -1), 'CENTER'),  # Data, Dia, Horário centralizados
            ('ALIGN', (3, 1), (-1, -1), 'LEFT'),   # Resto à esquerda
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e40af')),
            
            # Alternância de cores
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        elements.append(Paragraph(
            f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} - Sistema Apostello",
            footer_style
        ))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # ==================== ESCALA EXCEL ====================
    def gerar_escala_excel(self, escala_id: int, igreja_id: Optional[int] = None) -> BytesIO:
        """
        Gera Excel da escala mensal.
        Se igreja_id for informado, gera apenas os cultos dessa igreja.
        """
        escala = self.db.query(Escala).filter(Escala.id == escala_id).first()
        if not escala:
            raise ValueError("Escala não encontrada")
        
        distrito = self.db.query(Distrito).filter(Distrito.id == escala.distrito_id).first()
        
        # Filtrar itens
        query = self.db.query(ItemEscala).filter(ItemEscala.escala_id == escala_id)
        if igreja_id:
            query = query.filter(ItemEscala.igreja_id == igreja_id)
            # Se filtrado por igreja, ordenar por data/horário
            itens = query.order_by(ItemEscala.data_culto, ItemEscala.horario).all()
        else:
            # Se geral, ordenar por igreja primeiro, depois data/horário
            itens = query.order_by(ItemEscala.igreja_id, ItemEscala.data_culto, ItemEscala.horario).all()
        
        # Buscar igreja se filtrado
        igreja_filtro = None
        if igreja_id:
            igreja_filtro = self.db.query(Igreja).filter(Igreja.id == igreja_id).first()

        wb = Workbook()
        ws = wb.active
        
        if igreja_id and igreja_filtro:
            ws.title = f"{igreja_filtro.nome[:20]} {escala.mes:02d}-{escala.ano}"
        else:
            ws.title = f"Escala {escala.mes:02d}-{escala.ano}"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        # Traduzir status para português
        status_map = {
            'RASCUNHO': 'Rascunho',
            'PUBLICADA': 'Publicada',
            'ARQUIVADA': 'Arquivada'
        }
        status_texto = status_map.get(escala.status.value if hasattr(escala.status, 'value') else str(escala.status), str(escala.status))
        
        if igreja_id and igreja_filtro:
            # Cabeçalho para igreja específica
            ws.merge_cells('A1:F1')
            ws['A1'] = f"Escala de Pregação e Louvor - {igreja_filtro.nome}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")

            ws.merge_cells('A2:F2')
            ws['A2'] = f"{escala.mes:02d}/{escala.ano} - Distrito: {distrito.nome if distrito else 'N/A'} - Status: {status_texto}"
            ws['A2'].alignment = Alignment(horizontal="center")

            # Cabeçalhos (sem coluna Igreja)
            headers = ["Data", "Dia", "Horário", "Pregador", "Cantor", "Tema"]
        else:
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Escala de Pregação e Louvor - {escala.mes:02d}/{escala.ano}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")

            ws.merge_cells('A2:G2')
            ws['A2'] = f"Distrito: {distrito.nome if distrito else 'N/A'} - Status: {status_texto}"
            ws['A2'].alignment = Alignment(horizontal="center")

            # Cabeçalhos
            headers = ["Data", "Dia", "Horário", "Igreja", "Pregador", "Cantor", "Tema"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Mapeamento de dias da semana em português
        dias_semana_pt = {
            0: "Seg",
            1: "Ter",
            2: "Qua",
            3: "Qui",
            4: "Sex",
            5: "Sáb",
            6: "Dom"
        }

        # Dados
        for row_num, item in enumerate(itens, 5):
            pregador = self.db.query(Usuario).filter(Usuario.id == item.pregador_id).first() if item.pregador_id else None
            cantor = self.db.query(Usuario).filter(Usuario.id == item.cantor_id).first() if item.cantor_id else None
            
            dia_semana = dias_semana_pt.get(item.data_culto.weekday(), "") if item.data_culto else ""

            if igreja_id:
                # Sem coluna Igreja
                dados = [
                    item.data_culto.strftime("%d/%m/%Y") if item.data_culto else "",
                    dia_semana,
                    item.horario.strftime("%H:%M") if item.horario else "",
                    pregador.nome_completo if pregador else "A definir",
                    cantor.nome_completo if cantor else "-",
                    item.tema_customizado or (item.tema.titulo if item.tema else "-")
                ]
            else:
                igreja = self.db.query(Igreja).filter(Igreja.id == item.igreja_id).first()
                dados = [
                    item.data_culto.strftime("%d/%m/%Y") if item.data_culto else "",
                    dia_semana,
                    item.horario.strftime("%H:%M") if item.horario else "",
                    igreja.nome if igreja else "",
                    pregador.nome_completo if pregador else "A definir",
                    cantor.nome_completo if cantor else "-",
                    item.tema_customizado or (item.tema.titulo if item.tema else "-")
                ]

            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col, value=valor)
                cell.border = thin_border
                if col <= 3:
                    cell.alignment = Alignment(horizontal="center")

        # Ajustar largura das colunas
        if igreja_id:
            column_widths = [12, 8, 10, 30, 30, 30]
        else:
            column_widths = [12, 8, 10, 25, 30, 30, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ==================== PARTICIPAÇÕES ====================
    def gerar_participacoes_pdf(
        self, 
        distrito_id: int, 
        data_inicio: date, 
        data_fim: date
    ) -> BytesIO:
        """Gera relatório de participações por pregador/cantor"""
        # Buscar itens de escala no período
        itens = self.db.query(ItemEscala).join(
            Escala, ItemEscala.escala_id == Escala.id
        ).filter(
            Escala.distrito_id == distrito_id,
            ItemEscala.data_culto >= data_inicio,
            ItemEscala.data_culto <= data_fim
        ).all()

        # Contabilizar participações
        participacoes_pregador = {}
        participacoes_cantor = {}

        for item in itens:
            if item.pregador_id:
                if item.pregador_id not in participacoes_pregador:
                    pregador = self.db.query(Usuario).filter(Usuario.id == item.pregador_id).first()
                    participacoes_pregador[item.pregador_id] = {
                        "nome": pregador.nome_completo if pregador else "N/A",
                        "total": 0,
                        "sabados": 0,
                        "domingos": 0,
                        "outros": 0
                    }
                participacoes_pregador[item.pregador_id]["total"] += 1
                dia_semana = item.data_culto.weekday()
                if dia_semana == 5:
                    participacoes_pregador[item.pregador_id]["sabados"] += 1
                elif dia_semana == 6:
                    participacoes_pregador[item.pregador_id]["domingos"] += 1
                else:
                    participacoes_pregador[item.pregador_id]["outros"] += 1

            if item.cantor_id:
                if item.cantor_id not in participacoes_cantor:
                    cantor = self.db.query(Usuario).filter(Usuario.id == item.cantor_id).first()
                    participacoes_cantor[item.cantor_id] = {
                        "nome": cantor.nome_completo if cantor else "N/A",
                        "total": 0
                    }
                participacoes_cantor[item.cantor_id]["total"] += 1

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        elements.append(Paragraph(
            f"Relatório de Participações",
            title_style
        ))
        elements.append(Paragraph(
            f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
            ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, spaceAfter=20)
        ))

        # Tabela de Pregadores
        elements.append(Paragraph("Pregadores", styles['Heading2']))
        data_pregadores = [["Nome", "Total", "Sábados", "Domingos", "Outros"]]
        for p in sorted(participacoes_pregador.values(), key=lambda x: x["total"], reverse=True):
            data_pregadores.append([
                p["nome"], 
                str(p["total"]), 
                str(p["sabados"]), 
                str(p["domingos"]), 
                str(p["outros"])
            ])

        if len(data_pregadores) > 1:
            table = Table(data_pregadores)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Nenhum pregador encontrado no período.", styles['Normal']))

        elements.append(Spacer(1, 20))

        # Tabela de Cantores
        elements.append(Paragraph("Cantores", styles['Heading2']))
        data_cantores = [["Nome", "Total de Participações"]]
        for c in sorted(participacoes_cantor.values(), key=lambda x: x["total"], reverse=True):
            data_cantores.append([c["nome"], str(c["total"])])

        if len(data_cantores) > 1:
            table = Table(data_cantores)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Nenhum cantor encontrado no período.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def gerar_participacoes_excel(
        self, 
        distrito_id: int, 
        data_inicio: date, 
        data_fim: date
    ) -> BytesIO:
        """Gera Excel de participações"""
        # Mesma lógica do PDF para coletar dados
        itens = self.db.query(ItemEscala).join(
            Escala, ItemEscala.escala_id == Escala.id
        ).filter(
            Escala.distrito_id == distrito_id,
            ItemEscala.data_culto >= data_inicio,
            ItemEscala.data_culto <= data_fim
        ).all()

        participacoes_pregador = {}
        participacoes_cantor = {}

        for item in itens:
            if item.pregador_id:
                if item.pregador_id not in participacoes_pregador:
                    pregador = self.db.query(Usuario).filter(Usuario.id == item.pregador_id).first()
                    participacoes_pregador[item.pregador_id] = {
                        "nome": pregador.nome_completo if pregador else "N/A",
                        "total": 0,
                        "sabados": 0,
                        "domingos": 0,
                        "outros": 0
                    }
                participacoes_pregador[item.pregador_id]["total"] += 1
                dia_semana = item.data_culto.weekday()
                if dia_semana == 5:
                    participacoes_pregador[item.pregador_id]["sabados"] += 1
                elif dia_semana == 6:
                    participacoes_pregador[item.pregador_id]["domingos"] += 1
                else:
                    participacoes_pregador[item.pregador_id]["outros"] += 1

            if item.cantor_id:
                if item.cantor_id not in participacoes_cantor:
                    cantor = self.db.query(Usuario).filter(Usuario.id == item.cantor_id).first()
                    participacoes_cantor[item.cantor_id] = {
                        "nome": cantor.nome_completo if cantor else "N/A",
                        "total": 0
                    }
                participacoes_cantor[item.cantor_id]["total"] += 1

        wb = Workbook()
        
        # Aba Pregadores
        ws1 = wb.active
        ws1.title = "Pregadores"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        headers = ["Nome", "Total", "Sábados", "Domingos", "Outros"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, p in enumerate(sorted(participacoes_pregador.values(), key=lambda x: x["total"], reverse=True), 2):
            ws1.cell(row=row, column=1, value=p["nome"])
            ws1.cell(row=row, column=2, value=p["total"])
            ws1.cell(row=row, column=3, value=p["sabados"])
            ws1.cell(row=row, column=4, value=p["domingos"])
            ws1.cell(row=row, column=5, value=p["outros"])

        # Aba Cantores
        ws2 = wb.create_sheet("Cantores")
        headers = ["Nome", "Total de Participações"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, c in enumerate(sorted(participacoes_cantor.values(), key=lambda x: x["total"], reverse=True), 2):
            ws2.cell(row=row, column=1, value=c["nome"])
            ws2.cell(row=row, column=2, value=c["total"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ==================== AVALIAÇÕES ====================
    def gerar_avaliacoes_pdf(
        self, 
        distrito_id: int, 
        data_inicio: date, 
        data_fim: date
    ) -> BytesIO:
        """Gera relatório de avaliações"""
        avaliacoes = self.db.query(Avaliacao).join(
            Usuario, Avaliacao.avaliado_id == Usuario.id
        ).filter(
            Usuario.distrito_id == distrito_id,
            Avaliacao.created_at >= datetime.combine(data_inicio, datetime.min.time()),
            Avaliacao.created_at <= datetime.combine(data_fim, datetime.max.time())
        ).all()

        # Agrupar por avaliado
        por_avaliado = {}
        for av in avaliacoes:
            if av.avaliado_id not in por_avaliado:
                avaliado = self.db.query(Usuario).filter(Usuario.id == av.avaliado_id).first()
                por_avaliado[av.avaliado_id] = {
                    "nome": avaliado.nome_completo if avaliado else "N/A",
                    "tipo": avaliado.tipo.value if avaliado else "N/A",
                    "avaliacoes": [],
                    "media": 0
                }
            por_avaliado[av.avaliado_id]["avaliacoes"].append(av)

        # Calcular médias
        for uid, dados in por_avaliado.items():
            if dados["avaliacoes"]:
                total = sum(a.nota for a in dados["avaliacoes"])
                dados["media"] = total / len(dados["avaliacoes"])

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        elements.append(Paragraph(
            "Relatório de Avaliações",
            ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        ))
        elements.append(Paragraph(
            f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
            ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, spaceAfter=20)
        ))

        # Tabela resumo
        data_table = [["Nome", "Tipo", "Qtd Avaliações", "Média"]]
        for dados in sorted(por_avaliado.values(), key=lambda x: x["media"], reverse=True):
            data_table.append([
                dados["nome"],
                dados["tipo"],
                str(len(dados["avaliacoes"])),
                f"{dados['media']:.1f}"
            ])

        if len(data_table) > 1:
            table = Table(data_table)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Nenhuma avaliação encontrada no período.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def gerar_avaliacoes_excel(
        self, 
        distrito_id: int, 
        data_inicio: date, 
        data_fim: date
    ) -> BytesIO:
        """Gera Excel de avaliações"""
        avaliacoes = self.db.query(Avaliacao).join(
            Usuario, Avaliacao.avaliado_id == Usuario.id
        ).filter(
            Usuario.distrito_id == distrito_id,
            Avaliacao.created_at >= datetime.combine(data_inicio, datetime.min.time()),
            Avaliacao.created_at <= datetime.combine(data_fim, datetime.max.time())
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Avaliações"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        headers = ["Data", "Avaliado", "Tipo", "Avaliador", "Nota", "Comentário"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, av in enumerate(avaliacoes, 2):
            avaliado = self.db.query(Usuario).filter(Usuario.id == av.avaliado_id).first()
            avaliador = self.db.query(Usuario).filter(Usuario.id == av.avaliador_id).first()

            ws.cell(row=row, column=1, value=av.created_at.strftime("%d/%m/%Y") if av.created_at else "")
            ws.cell(row=row, column=2, value=avaliado.nome_completo if avaliado else "")
            ws.cell(row=row, column=3, value=avaliado.tipo.value if avaliado else "")
            ws.cell(row=row, column=4, value=avaliador.nome_completo if avaliador else "")
            ws.cell(row=row, column=5, value=av.nota)
            ws.cell(row=row, column=6, value=av.comentario or "")

        # Ajustar larguras
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
